"""Cria a aba DISCOS na familias.xlsx com as linhas das 3 famílias de disco
avulso (Composite, Disco de aço revestido por composite, Disco de aço), pra o
Leonardo rever a classificação separado.

Espelha as colunas A-G da aba CORREÇÃO exatamente (pra o merge de volta ser
fiel) e acrescenta a coluna I `linha_origem` = nº da linha na CORREÇÃO. Essa
coluna é referência técnica do merge — NÃO editar.

Idempotente: recria a aba DISCOS se já existir.
"""
from __future__ import annotations

import os
import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

# Caminho da planilha de famílias (rede). Vem do ambiente pra não fixar caminho
# nem identificação do cliente no repositório: export FAMILIAS_XLSX=...
_env = os.environ.get("FAMILIAS_XLSX")
if not _env:
    raise SystemExit("defina a variável de ambiente FAMILIAS_XLSX com o caminho da familias.xlsx")
PLANILHA = Path(_env)
FAMILIAS_DISCO = {"Composite", "Disco de aço revestido por composite", "Disco de aço"}
ABA = "DISCOS"


def main() -> None:
    backup = Path(__file__).resolve().parents[1] / "data" / f"familias.bak-{date.today().isoformat()}.xlsx"
    shutil.copy2(PLANILHA, backup)
    print(f"backup: {backup}")

    # Lê a planilha viva (sempre só leitura). Para ESCREVER, se o arquivo estiver
    # aberto no Excel (lock), grava num arquivo separado na mesma pasta.
    leitura = openpyxl.load_workbook(PLANILHA, data_only=True)
    src = leitura["CORREÇÃO"]

    travado = (PLANILHA.parent / f"~${PLANILHA.name}").exists()
    if travado:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        destino = PLANILHA.parent / "familias_DISCOS_revisao.xlsx"
    else:
        wb = openpyxl.load_workbook(PLANILHA)  # preserva fórmulas das outras abas
        if ABA in wb.sheetnames:
            del wb[ABA]
        destino = PLANILHA
    dst = wb.create_sheet(ABA)

    # cabeçalho: espelha A-G da CORREÇÃO + coluna técnica
    header = [src.cell(row=1, column=c).value for c in range(1, 8)]
    for c, v in enumerate(header, start=1):
        dst.cell(row=1, column=c, value=v)
    dst.cell(row=1, column=9, value="linha_origem")
    for c in list(range(1, 8)) + [9]:
        dst.cell(row=1, column=c).font = Font(bold=True)

    n = 0
    out_row = 2
    for i in range(2, src.max_row + 1):
        fam = src.cell(row=i, column=2).value
        if (str(fam).strip() if fam is not None else "") not in FAMILIAS_DISCO:
            continue
        for c in range(1, 8):
            dst.cell(row=out_row, column=c, value=src.cell(row=i, column=c).value)
        dst.cell(row=out_row, column=9, value=i)
        out_row += 1
        n += 1

    dst.freeze_panes = "A2"
    larguras = {"A": 18, "B": 38, "C": 16, "D": 12, "E": 10, "G": 52, "I": 12}
    for col, w in larguras.items():
        dst.column_dimensions[col].width = w

    wb.save(destino)
    onde = "arquivo separado (planilha estava aberta no Excel)" if travado else "na própria familias.xlsx"
    print(f"aba {ABA!r} com {n} linhas de disco (+ cabeçalho) -> {destino} [{onde}]")


if __name__ == "__main__":
    main()
