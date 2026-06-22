"""Merge das correções de família da aba DISCOS de volta na CORREÇÃO + alinha NCM.

A aba DISCOS foi reordenada pelo Leonardo e a coluna `linha_origem` desalinhou —
então re-chaveamos por (código, descrição), com fallback por código. Regras
combinadas com ele:
  - família VAZIA na DISCOS = "quero que vejam" -> IGNORA (não toca).
  - para família preenchida (verificada), o NCM EFETIVO na CORREÇÃO passa a ser o
    canônico da família (Composite / revestido -> 68138910; Disco de aço ->
    87084090), mesmo que ele só tenha mexido a família e não o NCM.
Convenção da CORREÇÃO preservada: col C = NCM atual, col D = "OK" (confirma C) ou
NCM corrigido. Para fixar efetivo=alvo: se alvo==C -> D="OK", senão D=alvo.
Banco (1 registro/código): ncm = canônico da família.
"""
from __future__ import annotations

import os
import shutil
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from src import produto  # noqa: E402

# Caminho da planilha de famílias (rede). Vem do ambiente pra não fixar caminho
# nem identificação do cliente no repositório: export FAMILIAS_XLSX=...
_env = os.environ.get("FAMILIAS_XLSX")
if not _env:
    raise SystemExit("defina a variável de ambiente FAMILIAS_XLSX com o caminho da familias.xlsx")
PLANILHA = Path(_env)
CANON = {
    "Composite": "68138910",
    "Disco de aço revestido por composite": "68138910",
    "Disco de aço": "87084090",
}


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def main() -> None:
    if (PLANILHA.parent / f"~${PLANILHA.name}").exists():
        print("ABORTADO: familias.xlsx está aberta no Excel. Feche e rode de novo.")
        return

    backup = Path(__file__).resolve().parents[1] / "data" / f"familias.bak-{date.today().isoformat()}-merge.xlsx"
    shutil.copy2(PLANILHA, backup)
    print(f"backup: {backup}")

    wb = openpyxl.load_workbook(PLANILHA)  # data_only=False: preserva fórmulas/abas
    d = wb["DISCOS"]
    o = wb["CORREÇÃO"]

    # índice da CORREÇÃO por (código, descrição) e por código (fallback)
    by_cd: dict = defaultdict(list)
    by_c: dict = defaultdict(list)
    for i in range(2, o.max_row + 1):
        cod = _s(o.cell(i, 1).value)
        if cod:
            by_cd[(cod, _s(o.cell(i, 7).value))].append(i)
            by_c[cod].append(i)

    ignoradas: list[str] = []
    fam_set: list[tuple] = []
    ncm_set: list[tuple] = []
    db_changes: dict[str, str] = {}
    usados: set[int] = set()

    for r in d.iter_rows(min_row=2, max_col=7, values_only=True):
        cod, fam, desc = _s(r[0]), _s(r[1]), _s(r[6])
        if not fam:
            ignoradas.append(cod)
            continue
        # casa a linha exata da CORREÇÃO (descrição desempata dups; evita reusar a mesma)
        cands = [i for i in by_cd.get((cod, desc), []) if i not in usados] or \
                [i for i in by_c.get(cod, []) if i not in usados]
        if not cands:
            print(f"  SEM MATCH: {cod} | {desc}")
            continue
        li = cands[0]
        usados.add(li)
        alvo = CANON[fam]

        if _s(o.cell(li, 2).value) != fam:
            fam_set.append((cod, _s(o.cell(li, 2).value), fam, li))
            o.cell(li, 2, value=fam)

        c_atual = _s(o.cell(li, 3).value)
        d_atual = _s(o.cell(li, 4).value)
        efetivo = c_atual if d_atual.upper() == "OK" else d_atual
        if efetivo != alvo:
            ncm_set.append((cod, efetivo or "(vazio)", alvo, li))
            if c_atual == alvo:
                o.cell(li, 4, value="OK")
            else:
                o.cell(li, 4, value=int(alvo))
            db_changes[cod] = alvo

    wb.save(PLANILHA)

    # banco
    db_aplicadas = []
    for cod, ncm in db_changes.items():
        p = produto.obter(cod)
        if p and p.ncm != ncm:
            produto.atualizar_campos(cod, {"ncm": ncm})
            db_aplicadas.append((cod, p.ncm, ncm))

    print(f"\nfamília corrigida na CORREÇÃO: {len(fam_set)}")
    for x in fam_set:
        print("   ", x)
    print(f"NCM efetivo alinhado na CORREÇÃO: {len(ncm_set)}")
    for x in ncm_set:
        print("   ", x)
    print(f"banco alterado: {len(db_aplicadas)}")
    for x in db_aplicadas:
        print("   ", x)
    print(f"família vazia ignorada: {len(ignoradas)} -> {ignoradas}")


if __name__ == "__main__":
    main()
