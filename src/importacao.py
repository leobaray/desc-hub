"""Importador de planilha -> cadastro (Produto).

O caminho do "mistura": você exporta a planilha completa (já pré-preenchida com o
que o sistema sabe), preenche os campos manuais no Excel e reimporta aqui. Cada
linha faz upsert no cadastro PELO CÓDIGO: preenche o que veio, preserva o resto
(célula vazia não apaga dado já cadastrado).

Tolerante a cabeçalho: normaliza o nome da coluna (sem acento, sem pontuação,
minúsculo) e casa com o campo do Produto por uma tabela de apelidos. Use o layout
da exportação completa; a coluna "Imagem" é ignorada (imagem entra por upload, não
pela célula).

    python -m src.importacao caminho/da/planilha.xlsx
"""
from __future__ import annotations

import io
import re
import unicodedata
from pathlib import Path

from src import produto as produto_mod
from src.saida import COLUNAS_COMPLETA


def _norm(s: object) -> str:
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())


# cabeçalho normalizado -> campo do Produto. Base: o layout da exportação completa.
_MAPA: dict[str, str] = {}
for _cab, _campo in COLUNAS_COMPLETA:
    if _campo and not _campo.startswith("_"):
        _MAPA[_norm(_cab)] = _campo

# Apelidos extras (variações de nome que sua planilha pode ter).
_APELIDOS = {
    "codigo": "codigo", "codprod": "codigo", "codigodoproduto": "codigo",
    "codinterno": "cod_ss", "codigointerno": "cod_ss",
    "codcatalogodeprodutosiscomex": "cod_sisc", "codigosiscomex": "cod_sisc",
    "codsiscomex": "cod_sisc",
    "codigocclasstrib": "cclasstrib", "classtrib": "cclasstrib",
    "pais": "pais_origem", "paisorigem": "pais_origem",
    "nvemateriaprima": "nve_materia_prima",
    "nveprocesso": "nve_processo", "nveprocessofabricacao": "nve_processo",
    "nveacabamento": "nve_acabamento",
    "quantiadaembalagemnaentrada": "qtd_embalagem_entrada",
    "quantiadaembalagemnasaida": "qtd_embalagem_saida",
    "unidadedemedidadaembalagemnasaida": "un_medida_saida",
    "unmedidaentrada": "un_medida_entrada",
    "localizacaonoestoque": "localizacao_estoque", "localizacao": "localizacao_estoque",
    "aplicacoesinformacoesdosdadostecnicos": "aplicacoes", "aplicacao": "aplicacoes",
    "revenda": "revenda_uso_interno", "usointerno": "revenda_uso_interno",
    "revendausointerno": "revenda_uso_interno",
    "descricaogeral": "descricao", "descricaocomercial": "descricao",
}
_MAPA.update(_APELIDOS)

# nunca importar pela célula (vêm de upload/sistema)
_IGNORAR = {"imagem"}


def _mapear_cabecalhos(cabecalhos: list) -> dict[int, str]:
    """índice da coluna -> campo do Produto (só as que reconhecemos)."""
    out: dict[int, str] = {}
    for i, cab in enumerate(cabecalhos):
        campo = _MAPA.get(_norm(cab))
        if campo and campo not in _IGNORAR:
            out[i] = campo
    return out


def importar(origem: str | Path | bytes) -> dict:
    """Lê o xlsx e faz upsert linha a linha. Devolve um resumo."""
    from openpyxl import load_workbook

    fonte = io.BytesIO(origem) if isinstance(origem, bytes) else origem
    wb = load_workbook(fonte, read_only=True, data_only=True)
    ws = wb.active

    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return {"erro": "planilha vazia", "criados": 0, "atualizados": 0, "linhas": 0}

    col2campo = _mapear_cabecalhos(list(linhas[0]))
    if "codigo" not in col2campo.values():
        return {
            "erro": "não achei a coluna do código (ex.: CódigoFabricante)",
            "reconhecidas": sorted(set(col2campo.values())),
            "criados": 0, "atualizados": 0, "linhas": 0,
        }

    criados = atualizados = ignoradas = 0
    campos_vistos: set[str] = set()
    for raw in linhas[1:]:
        valores = {
            col2campo[i]: ("" if v is None else str(v).strip())
            for i, v in enumerate(raw)
            if i in col2campo
        }
        codigo = (valores.pop("codigo", "") or "").strip()
        if not codigo:
            ignoradas += 1
            continue
        # só campos preenchidos: célula vazia preserva o que já existe no cadastro
        campos = {k: v for k, v in valores.items() if v != ""}
        existia = produto_mod.existe(codigo)
        produto_mod.atualizar_campos(codigo, campos)
        campos_vistos.update(campos.keys())
        if existia:
            atualizados += 1
        else:
            criados += 1

    return {
        "criados": criados,
        "atualizados": atualizados,
        "ignoradas_sem_codigo": ignoradas,
        "linhas": len(linhas) - 1,
        "campos_preenchidos": sorted(campos_vistos),
    }


# ---------------------------------------------------------------------------
# Arquivo-mestre de famílias (todos os códigos da empresa)
# ---------------------------------------------------------------------------
# Layout da aba "familias":
#   A codigo_fabricante | B familia | C classif_fiscal_ncm | D marca |
#   E <status: ok / ? / Descontinuado> | F descricao_do_produto
#
# Por ora importamos SÓ o código + o fabricante (marca). Família, NCM e descrição
# ainda não estão fechados (a contabilidade vai renomear/fundir famílias e arrumar
# os NCM) — entram quando o arquivo final voltar, ligando as flags abaixo.

# Marcas que NÃO viram fabricante (são genéricas / "outros").
_MARCAS_GENERICAS = {
    "DIVER", "DIVERSOS", "DIVERSO", "OUTROS", "OUTRO", "VARIOS", "VARIAS",
    "NA", "N/A", "-", "?", "SEM", "SEMMARCA", "INDEFINIDO",
}


def _eh_marca_generica(marca: str) -> bool:
    return marca.strip().upper() in _MARCAS_GENERICAS


# Sigla do arquivo -> nome canônico do fabricante (confirmados pelo Leonardo).
_MARCA_CANONICA = {
    "RAYBE": "Raybestos",
    "RAYBESTOS": "Raybestos",
    "ALLOM": "Allomatic",
    "SONNA": "Sonnax",
    "ALTO": "Alto",
    "TRICO": "Tricomponent",
    "PSMFG": "PS Bearings",
}


def _canonizar_marca(marca: str) -> str:
    return _MARCA_CANONICA.get(marca.strip().upper(), marca.strip())


def importar_master_familias(
    origem: str | Path | bytes,
    *,
    trazer_marca: bool = True,
    trazer_descricao: bool = False,
    trazer_ncm: bool = False,
    trazer_familia: bool = False,
) -> dict:
    """Importa o arquivo-mestre 'familias.xlsx' pro cadastro.

    Upsert por código, sempre **seed-if-empty**: nunca sobrescreve campo já
    preenchido (protege os 95 já migrados e qualquer edição manual). Marca
    genérica (DIVER/"outros") não vira fabricante. Conflitos de NCM entre linhas
    do mesmo código são reportados (não corrigidos)."""
    from openpyxl import load_workbook

    fonte = io.BytesIO(origem) if isinstance(origem, bytes) else origem
    wb = load_workbook(fonte, read_only=True, data_only=True)
    ws = wb["familias"] if "familias" in wb.sheetnames else wb.active

    linhas = novos = existentes = sem_codigo = 0
    fabricante_setado = marca_generica = 0
    vistos: set[str] = set()
    ncm_por_codigo: dict[str, str] = {}
    conflitos_ncm: list[dict] = []

    for raw in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        codigo = str(raw[0]).strip() if raw[0] is not None else ""
        if not codigo:
            sem_codigo += 1
            continue
        linhas += 1
        familia = str(raw[1]).strip() if len(raw) > 1 and raw[1] is not None else ""
        ncm = str(raw[2]).strip() if len(raw) > 2 and raw[2] is not None else ""
        marca = str(raw[3]).strip() if len(raw) > 3 and raw[3] is not None else ""
        descricao = str(raw[5]).strip() if len(raw) > 5 and raw[5] is not None else ""

        # conflito de NCM no mesmo código (só pra avisar)
        if ncm:
            anterior = ncm_por_codigo.get(codigo)
            if anterior and anterior != ncm:
                conflitos_ncm.append({"codigo": codigo, "ncm_a": anterior, "ncm_b": ncm})
            ncm_por_codigo.setdefault(codigo, ncm)

        # contagem novo/existente: só na 1ª vez que o código aparece neste run
        if codigo not in vistos:
            if produto_mod.existe(codigo):
                existentes += 1
            else:
                novos += 1
            vistos.add(codigo)

        p = produto_mod.obter(codigo) or produto_mod.Produto(codigo=codigo)

        if trazer_marca and marca and not (p.fabricante or "").strip():
            if _eh_marca_generica(marca):
                marca_generica += 1
            else:
                p.fabricante = _canonizar_marca(marca)
                fabricante_setado += 1
        if trazer_descricao and descricao and not (p.descricao or "").strip():
            p.descricao = descricao
        if trazer_ncm and ncm and not (p.ncm or "").strip():
            p.ncm = ncm
        if trazer_familia and familia and hasattr(p, "familia") and not (getattr(p, "familia", "") or "").strip():
            setattr(p, "familia", familia)

        produto_mod.salvar(p)

    wb.close()
    return {
        "linhas_com_codigo": linhas,
        "codigos_distintos": len(vistos),
        "novos": novos,
        "ja_existiam": existentes,
        "sem_codigo": sem_codigo,
        "fabricante_setado": fabricante_setado,
        "marca_generica_puladas": marca_generica,
        "conflitos_ncm": conflitos_ncm,
    }


def normalizar_fabricantes() -> dict:
    """Reescreve os fabricantes já cadastrados pra forma canônica (RAYBE e
    raybestos -> Raybestos, SONNA -> Sonnax, ...). Não toca em marca sem
    mapeamento (TRICO, PSMFG) nem em fabricante vazio."""
    alterados = 0
    antes_depois: dict[str, str] = {}
    for p in produto_mod.listar():  # listar() já inicializa o banco
        atual = (p.fabricante or "").strip()
        if not atual:
            continue
        canon = _canonizar_marca(atual)
        if canon != atual:
            produto_mod.atualizar_campos(p.codigo, {"fabricante": canon})
            alterados += 1
            antes_depois[atual] = canon
    return {"alterados": alterados, "mapeamentos": antes_depois}


if __name__ == "__main__":
    import sys

    args = sys.argv[1:]
    if args and args[0] == "--normalizar-fabricantes":
        print(normalizar_fabricantes())
    elif args and args[0] == "--master":
        if len(args) < 2:
            print("uso: python -m src.importacao --master caminho/familias.xlsx")
            raise SystemExit(2)
        print(importar_master_familias(args[1]))
    elif args:
        print(importar(args[0]))
    else:
        print("uso: python -m src.importacao caminho/planilha.xlsx")
        print("  ou: python -m src.importacao --master caminho/familias.xlsx")
        raise SystemExit(2)
